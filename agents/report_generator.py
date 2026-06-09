import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from config import REPORTS_DIR
from utils.helpers import format_currency


class ReportGenerator:
    def __init__(self):
        self.styles = {
            "header": Font(bold=True, color="FFFFFF"),
            "title": Font(bold=True, size=16, color="1F2D3D"),
            "subtitle": Font(bold=True, size=12, color="1F2D3D"),
            "fill_header": PatternFill("solid", fgColor="1F4E79"),
            "fill_alt": PatternFill("solid", fgColor="F2F2F2"),
            "center": Alignment(horizontal="center", vertical="center"),
            "wrap": Alignment(wrap_text=True, vertical="top")
        }

    def generate_report(self, analysis_results: Dict[str, Any], output_path: str, summary: str = ""):
        """Gera relatório Excel com dados, gráficos e análise."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo"

        # --- Sheet 1: Resumo Executivo ---
        self._write_header(ws, "RESUMO EXECUTIVO", row=1)
        ws.merge_cells(f"A1:D1")
        ws["A2"] = summary
        ws.merge_cells(f"A2:D4")
        ws["A2"].alignment = self.styles["wrap"]

        # Insights
        ws["A6"] = "🔍 Principais Insights"
        ws["A6"].font = self.styles["subtitle"]
        insights = analysis_results.get("insights", [])
        for i, insight in enumerate(insights, start=7):
            ws[f"A{i}"] = f"• {insight}"
            ws[f"A{i}"].alignment = self.styles["wrap"]

        # Estatísticas gerais
        ws["F6"] = "📊 Estatísticas Gerais"
        ws["F6"].font = self.styles["subtitle"]
        ws["F7"] = f"Total de produtos analisados: {analysis_results.get('total_products_analyzed', 0)}"
        ws["F8"] = f"Janela de análise: {analysis_results.get('analysis_window_days', 30)} dias"

        # --- Sheet 2: Top Produtos ---
        ws2 = wb.create_sheet("Top Produtos")
        headers = ["Rank", "Produto", "Categoria", "Plataforma", "Preço", "Avaliação", "Vendas", "Trend Score", "Competitividade", "Score Final", "Link"]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = self.styles["header"]
            cell.fill = self.styles["fill_header"]
            cell.alignment = self.styles["center"]

        top_products = analysis_results.get("top_products", [])
        for idx, p in enumerate(top_products, start=2):
            ws2[f"A{idx}"] = idx - 1
            ws2[f"B{idx}"] = p.get("title", "")[:80]
            ws2[f"C{idx}"] = p.get("category", "")
            ws2[f"D{idx}"] = p.get("platform", "")
            ws2[f"E{idx}"] = p.get("price_num", 0)
            ws2[f"F{idx}"] = p.get("rating_num", 0)
            ws2[f"G{idx}"] = p.get("sales_num", 0)
            ws2[f"H{idx}"] = round(p.get("trend_score", 0), 3)
            ws2[f"I{idx}"] = round(p.get("competitiveness", 0), 3)
            ws2[f"J{idx}"] = round(p.get("final_score", 0), 3)
            ws2[f"K{idx}"] = p.get("url", "")
            ws2[f"K{idx}"].hyperlink = p.get("url", "")

        # Formatar colunas
        col_widths = [5, 40, 15, 12, 10, 10, 8, 12, 14, 11, 25]
        for i, w in enumerate(col_widths, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # Formatar preços
        for r in range(2, 2 + len(top_products)):
            ws2[f"E{r}"].number_format = '"R$" #,##0.00'

        # Alternar cores nas linhas
        for r in range(2, 2 + len(top_products)):
            if r % 2 == 0:
                for c in range(1, 12):
                    ws2.cell(row=r, column=c).fill = self.styles["fill_alt"]

        # --- Sheet 3: Análise por Categoria ---
        ws3 = wb.create_sheet("Análise por Categoria")
        ws3.append(["Categoria", "Média de Preço", "Mediana", "Total de Produtos", "Score Médio"])
        for cell in ws3[1]:
            cell.font = self.styles["header"]
            cell.fill = self.styles["fill_header"]
            cell.alignment = self.styles["center"]

        price_stats = analysis_results.get("price_stats", {})
        by_cat = price_stats.get("by_category", {})
        top_by_cat = analysis_results.get("top_by_category", {})
        for i, (cat, stats) in enumerate(by_cat.items(), start=2):
            ws3[f"A{i}"] = cat
            ws3[f"B{i}"] = stats.get("mean", 0)
            ws3[f"C{i}"] = stats.get("median", 0)
            ws3[f"D{i}"] = stats.get("count", 0)
            # Score médio da categoria
            products = top_by_cat.get(cat, [])
            avg_score = np.mean([p.get("final_score", 0) for p in products]) if products else 0
            ws3[f"E{i}"] = round(avg_score, 3)
            ws3[f"B{i}"].number_format = '"R$" #,##0.00'

        # Gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.title = "Média de Preço por Categoria"
        chart.y_axis.title = "Preço Médio (R$)"
        chart.x_axis.title = "Categoria"

        data = Reference(ws3, min_col=2, min_row=1, max_row=1 + len(by_cat))
        cats = Reference(ws3, min_col=1, min_row=2, max_row=1 + len(by_cat))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 22
        ws3.add_chart(chart, "G2")

        # Salvar
        wb.save(output_path)

    def _write_header(self, ws, text: str, row: int = 1):
        ws[f"A{row}"] = text
        ws[f"A{row}"].font = self.styles["title"]

    def _format_currency(self, value: float) -> str:
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
